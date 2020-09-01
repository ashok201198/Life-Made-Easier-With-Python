import pymysql as pms
from MySQLdb import Error
from openpyxl import *
import warnings
warnings.filterwarnings("ignore")

conn = pms.connect(host="localhost", user="root", passwd="123456789")
cur1 = conn.cursor()  # prepare a cursor object using cursor() method
cur1.execute("DROP SCHEMA IF EXISTS summerdata")  ### Drop the DATABASE if it already exist ###

query = "CREATE SCHEMA `summerdata`;"
cur1.execute(query)
conn.select_db('summerdata')
""" Creating students and marks tables"""

query = """CREATE TABLE STUDENT(
                    NAME  CHAR(50) NOT NULL,
                    COLLEGE CHAR(50) NOT NULL,
                    EMAIL CHAR(50) NOT NULL,
                    DB_NAME CHAR(50) NOT NULL,
                    PRIMARY KEY(DB_NAME))"""
cur1.execute(query)

query = """CREATE TABLE MARKS (
            NAME  CHAR(40) NOT NULL,
            TRANSFORM INT NOT NULL,
            FROM_CUSTOM_BASE26 INT NOT NULL,
            GET_PIG_LATIN INT NOT NULL,
            TOP_CHARS INT NOT NULL,
            TOTAL INT NOT NULL,
            PRIMARY KEY(NAME),
            FOREIGN KEY (NAME) REFERENCES STUDENT(DB_NAME))"""
cur1.execute(query)
students_details=load_workbook('students1.xlsx')
ws=students_details['Current']
for row in range(2,ws.max_row+1):
    details=[]
    for col in range(1,ws.max_column+1):
        details.append(ws.cell(row=row,column=col).value)
    details[3]=details[3].lower()
    q="INSERT INTO summerdata.STUDENT(NAME,COLLEGE,EMAIL,DB_NAME) VALUES('%s','%s','%s','%s')"%(details[0],details[1],details[2],details[3])
    conn.query(q)
    print(details,"inserted")
conn.commit()
ws=students_details['Deletions']
for row in range(2,ws.max_row+1):
    details=[]
    for col in range(1,ws.max_column+1):
        details.append(ws.cell(row=row,column=col).value)
    details[3]=details[3].lower()
    q="INSERT INTO summerdata.STUDENT(NAME,COLLEGE,EMAIL,DB_NAME) VALUES('%s','%s','%s','%s')"%(details[0],details[1],details[2],details[3])
    conn.query(q)
    print(details,"inserted")
conn.commit()
students_details.close()
marks_details=load_workbook('marks2.xlsx')
ws=marks_details.active
print(ws)
cur1=conn.cursor()
for row in range(3,ws.max_row+1):
    details=[]
    for col in range(1,ws.max_column+1):
        details.append(ws.cell(row=row,column=col).value)
    st=details[0].split('_')
    s=st[2].lower()
    try:
        q = "INSERT INTO summerdata.MARKS VALUES('%s','%d','%d','%d','%d','%d')" % (s, details[1], details[2], details[3],details[4],details[5])
        cur1.execute(q)
    except Exception as e:
        pass
    print(s,details[1:],"inserted")
    conn.commit()
conn.commit()
conn.close()
#(NAME,TRANSFORM,FROM_CUSTOM_BASE26,GET_PIG_LATIN,TOP_CHARS,TOTAL)