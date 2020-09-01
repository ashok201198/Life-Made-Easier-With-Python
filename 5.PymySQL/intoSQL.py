import pymysql as db
import warnings
import click
from openpyxl import *
warnings.filterwarnings("ignore")

@click.group()
def cli():
    pass


@cli.command('createdb')
def createdb():
    try:
        conn = db.connect(host="localhost", user="root", passwd="123456789")
        cur1 = conn.cursor()  # prepare a cursor object using cursor() method

        cur1.execute("DROP SCHEMA IF EXISTS onlinecollege")  ### Drop the DATABASE if it already exist ###

        query = "CREATE SCHEMA `onlinecollege`;"
        cur1.execute(query)
        conn.select_db('onlinecollege')
        """ Creating students and marks tables"""

        query = """CREATE TABLE STUDENT(
                    NAME1  CHAR(50) NOT NULL,
                    COLLEGE CHAR(50) NOT NULL,
                    EMAIL CHAR(50) NOT NULL,
                    DB_NAME CHAR(50) NOT NULL,
                    PRIMARY KEY(DB_NAME))"""
        cur1.execute(query)

        query = """CREATE TABLE MARKS (
                        NAME2  CHAR(40) NOT NULL,
                        TRANSFORM INT NOT NULL,
                        FROM_CUSTOM_BASE26 INT NOT NULL,
                        GET_PIG_LATIN INT NOT NULL,
                        TOP_CHARS INT NOT NULL,
                        TOTAL INT NOT NULL,
                        PRIMARY KEY(NAME2),
                        FOREIGN KEY (NAME2) REFERENCES STUDENT(DB_NAME))"""
        cur1.execute(query)
        conn.close()
        click.echo( "Successfully created database and the tables students and marks")
    except Exception as e:
        print(e)

@cli.command('dropdb')
def dropdb():
    try:
        conn = db.connect(host="localhost", user="root", passwd="Ashok.123")
        cur1 = conn.cursor()  # prepare a cursor object using cursor() method

        cur1.execute("DROP SCHEMA IF EXISTS onlinecollege")  ### Drop the DATABASE if it already exist ###
        click.echo("Dropped Successfully")
    except Exception as e:
        print(e)

@cli.command('importdata')
def importdata():
    try:
        conn = db.connect(host="localhost", user="root", passwd="123456789")
        curs=conn.cursor()
        students_details=load_workbook('students1.xlsx')
        ws =students_details['Current']
        for row in range(2, ws.max_row + 1):
            details = []
            for col in range(1, ws.max_column + 1):
                details.append(ws.cell(row=row, column=col).value)
            details[3]="ol2016_"+details[1].lower()+"_"+details[3].lower()+"_mock"
            try:
                q = "INSERT INTO onlinecollege.STUDENT VALUES('%s','%s','%s','%s')" % (details[0], details[1], details[2],details[3])
                curs.execute(q)
            except Exception as e:
                pass
            conn.commit()
        conn.commit()
        ws = students_details['Deletions']
        for row in range(2, ws.max_row + 1):
            details = []
            for col in range(1, ws.max_column + 1):
                details.append(ws.cell(row=row, column=col).value)
            details[3] = "ol2016_"+details[1].lower()+"_"+details[3].lower()+"_mock"
            try:
                q = "INSERT INTO onlinecollege.STUDENT VALUES('%s','%s','%s','%s')" % (details[0], details[1], details[2], details[3])
                curs.execute(q)
            except Exception as e:
                pass
            conn.commit()
        conn.commit()
        students_details.close()
        marks_details = load_workbook('marks2.xlsx')
        ws = marks_details.active
        #print(ws)
        for row in range(3, ws.max_row + 1):
            details = []
            for col in range(1, ws.max_column + 1):
                details.append(ws.cell(row=row, column=col).value)
            try:
                q = "INSERT INTO onlinecollege.MARKS VALUES('%s','%d','%d','%d','%d','%d')" % (details[0], details[1], details[2], details[3], details[4], details[5])
                curs.execute(q)
            except Exception as e:
                print(e)
            conn.commit()
        conn.commit()
        conn.close()
        click.echo("Data imported from students.xlsx and marks.xlsx successfully")
    except Exception as e:
        click.echo(e)

@cli.command('collegestats')
def collegestats():
    try:
        conn = db.connect(host="localhost", user="root", passwd="Ashok.123")
        conn.select_db('onlinecollege')
        cur1=conn.cursor()

        # cur1.execute("SELECT * FROM student")
        # j=0
        # for i in cur1.fetchall():
        #     j+=1
        # print(j)
        #
        # cur1.execute("SELECT * FROM marks")
        # j=0
        # for i in cur1.fetchall():
        #     j+=1
        # print(j)

        q="select student.college,count(*),min(marks.total),max(marks.total),avg(marks.total) from student , marks where marks.NAME2 =student.DB_NAME group by student.college"
        cur1.execute(q)
        count=0
        for i in cur1.fetchall():
            click.echo(i)
            count+=1
        print(count)
    except Exception as e:
        print(e)


cli()