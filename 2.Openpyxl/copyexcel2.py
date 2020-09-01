import openpyxl
from openpyxl.styles import Font
x="students.xlsx"
wb=openpyxl.load_workbook(x)
ws=wb.active
print(wb.sheetnames)
for i in wb:
    for j in range(1,i.max_row+1):
        for k in range(1,i.max_column+1):
            if type(i.cell(row=j,column=k).value).__name__=="str":
                f=i.cell(row=j, column=k).font.__copy__()
                print(f)
                i.cell(row=j, column=k).value=i.cell(row=j,column=k).value.capitalize()
                i.cell(row=j, column=k).font=f
for i in wb:
    for j in range(1,i.max_row+1):
        for k in range(1,i.max_column+1):
            i.cell(row=j, column=k).font=Font()
wb.save("newone.xlsx")

#to save font use copy