import click
import openpyxl
import os
import datetime
from openpyxl.styles import Font
@click.command()
@click.option("--capitalize",is_flag=True,help="To capitalize all strings",default=False)
@click.option("--preservestyles",is_flag=True,help="To preserve styles of all strings",default=False)
@click.argument("source_excel")
@click.argument("dest_excel")
def cli(capitalize,preservestyles, source_excel,dest_excel):
    workbook = openpyxl.load_workbook(source_excel)
    if os.path.isfile(dest_excel):
        if not click.prompt("Do you want to override?",default=False):
            return
    #before=datetime.datetime.now()
    for worksheet in workbook:
        for row_no in range(1, worksheet.max_row + 1):
            for col_no in range(1, worksheet.max_column + 1):
                fonts_applied=worksheet.cell(row=row_no, column=col_no).font.__copy__()
                if capitalize:
                    if type(worksheet.cell(row=row_no, column=col_no).value).__name__ == "str":
                        worksheet.cell(row=row_no, column=col_no).value = worksheet.cell(row=row_no, column=col_no).value.capitalize()
                    else:
                        worksheet.cell(row=row_no, column=col_no).value = worksheet.cell(row=row_no, column=col_no).value
                if preservestyles:
                    worksheet.cell(row=row_no, column=col_no).font = fonts_applied
                else:
                    worksheet.cell(row=row_no, column=col_no).font = Font()
    #x=datetime.datetime.now()-before
    #print(x.seconds)
    #print(x.microseconds)
    workbook.save(dest_excel)
cli()