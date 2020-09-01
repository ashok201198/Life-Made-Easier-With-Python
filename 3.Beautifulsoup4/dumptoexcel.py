import click
from bs4 import BeautifulSoup
import openpyxl

@click.command()
@click.argument("source_html")
@click.argument("dest_excel")
def cli(source_html,dest_excel):
    fp = open(dest_excel, 'w')
    file = open(source_html)
    #file is opened because the html is a big file
    soup = BeautifulSoup(file,"html5lib")
    #we can use html5 parser also
    table_rows = soup.find_all('tr')
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    row_no = 1
    column_no = 1
    for table_row in table_rows[0].find_all('th')[1:]:
        worksheet.cell(row=row_no, column=column_no).value = table_row.text
        column_no += 1
    row_no, column_no = 2, 1
    for table_row in table_rows:
        for row_data in table_row.find_all('td')[1:]:
            if row_data.text.strip().isdigit():
                worksheet.cell(row=row_no, column=column_no).value = int(row_data.text)
            else:
                worksheet.cell(row=row_no, column=column_no).value = row_data.text
            column_no += 1
        column_no = 1
        row_no += 1
    dimensions = {}
    for row in worksheet.rows:
        for cell in row:
            dimensions[cell.column] = max((dimensions.get(cell.column, 0), len(str(cell.value))))
    for col, value in dimensions.items():
        worksheet.column_dimensions[col].width = value
    workbook.save(dest_excel)


cli()