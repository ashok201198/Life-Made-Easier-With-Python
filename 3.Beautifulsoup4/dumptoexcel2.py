from bs4 import BeautifulSoup
import openpyxl
x="sample.xlsx"
file=open(x,'w')
file=open("mockresults.html")
soup=BeautifulSoup(file,"html5lib")
list_all=soup.find_all('tr')
work1=openpyxl.Workbook()
work=work1.active
i=1
j=1
for l in list_all[0].find_all('th')[1:]:
    work.cell(row=i,column=j).value=l.text
    #print(l.text)
    j+=1
i,j=2,1
for l in list_all:
    for m in l.find_all('td')[1:]:
        work.cell(row=i, column=j).value = m.text
        j+=1
    j=1
    i+=1
ws = work
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
for col, value in dims.items():
    ws.column_dimensions[col].width = value
work1.save(x)